"""
询价/需求管理路由模块
"""
import io
import csv
import urllib.parse
import openpyxl
from openpyxl.styles import Font, Alignment
from fastapi import APIRouter, Request, Form, Depends, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse

from Sills.base import get_db_connection
from Sills.db_cli import get_cli_list
from Sills.db_quote import (
    get_quote_list, add_quote, batch_import_quote_text, batch_import_quote_from_rows,
    update_quote, delete_quote, batch_delete_quote, batch_copy_quote, batch_add_quotes
)
from routes.auth import login_required, get_current_user, templates

# 内部API Key（需要从main.py导入）
INTERNAL_API_KEY = None

router = APIRouter(prefix="/quote", tags=["quote"])


@router.get("", response_class=HTMLResponse)
async def quote_page(
    request: Request, current_user: dict = Depends(login_required),
    page: int = 1, page_size: int = 20, search: str = "",
    start_date: str = "", end_date: str = "", cli_id: str = "",
    status: str = "", is_transferred: str = ""
):
    """询价列表页面"""
    session = request.session
    has_params = any(k in request.query_params for k in ['search', 'start_date', 'end_date', 'cli_id', 'status', 'is_transferred'])

    if not has_params:
        search = session.get("quote_search", "")
        start_date = session.get("quote_start_date", "")
        end_date = session.get("quote_end_date", "")
        cli_id = session.get("quote_cli_id", "")
        status = session.get("quote_status", "")
        is_transferred = session.get("quote_is_transferred", "")
    else:
        session["quote_search"] = search
        session["quote_start_date"] = start_date
        session["quote_end_date"] = end_date
        session["quote_cli_id"] = cli_id
        session["quote_status"] = status
        session["quote_is_transferred"] = is_transferred

    results, total = get_quote_list(
        page=page, page_size=page_size, search_kw=search,
        start_date=start_date, end_date=end_date,
        cli_id=cli_id, status=status, is_transferred=is_transferred
    )
    total_pages = (total + page_size - 1) // page_size
    cli_list, _ = get_cli_list(page=1, page_size=1000)
    cli_list = sorted(cli_list, key=lambda x: x.get('cli_name', ''))

    return templates.TemplateResponse("quote.html", {
        "request": request,
        "active_page": "quote",
        "current_user": current_user,
        "items": results,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "search": search,
        "start_date": start_date,
        "end_date": end_date,
        "cli_id": cli_id,
        "status": status,
        "is_transferred": is_transferred,
        "cli_list": cli_list
    })


@router.post("/add")
async def quote_add(request: Request, current_user: dict = Depends(login_required)):
    """添加询价"""
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/quote", status_code=303)
    form = await request.form()
    data = dict(form)
    ok, msg = add_quote(data)
    msg_param = urllib.parse.quote(msg)
    success = 1 if ok else 0
    return RedirectResponse(url=f"/quote?msg={msg_param}&success={success}", status_code=303)


@router.post("/import")
async def quote_import_text(batch_text: str = Form(...), current_user: dict = Depends(login_required)):
    """批量导入询价（文本）"""
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/quote", status_code=303)
    success_count, errors = batch_import_quote_text(batch_text)
    err_msg = ""
    if errors:
        err_msg = "&msg=" + urllib.parse.quote(errors[0])
    return RedirectResponse(url=f"/quote?import_success={success_count}&errors={len(errors)}{err_msg}", status_code=303)


@router.get("/api/quote/template")
async def api_quote_template(current_user: dict = Depends(get_current_user)):
    """下载需求导入模板 (Excel格式)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "需求导入模板"

    headers = ["日期", "客户名", "询价型号", "报价型号", "询价品牌", "询价数量", "目标价", "成本价", "批号", "交期", "状态", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    ws.cell(row=2, column=1, value="2026-01-01")
    ws.cell(row=2, column=2, value="示例客户")
    ws.cell(row=2, column=3, value="STM32F103C8T6")
    ws.cell(row=2, column=4, value="STM32F103C8T6")
    ws.cell(row=2, column=5, value="ST")
    ws.cell(row=2, column=6, value=500)
    ws.cell(row=2, column=7, value=8.5)
    ws.cell(row=2, column=8, value=7.0)
    ws.cell(row=2, column=9, value="2912+")
    ws.cell(row=2, column=10, value="1~3days")
    ws.cell(row=2, column=11, value="询价中")
    ws.cell(row=2, column=12, value="示例数据")

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws.column_dimensions[col].width = 12 if col in ['A', 'E', 'F', 'G', 'H', 'I', 'J', 'K'] else 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=quote_template.xlsx"}
    )


@router.post("/import/csv")
async def quote_import_csv(csv_file: UploadFile = File(...), current_user: dict = Depends(login_required)):
    """批量导入询价（CSV/Excel）"""
    if current_user['rule'] not in ['3', '0']:
        return RedirectResponse(url="/quote", status_code=303)

    content = await csv_file.read()
    filename = csv_file.filename or ""
    rows_data = []

    if filename.lower().endswith('.xlsx'):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                rows_data.append(row_values)
        except Exception as e:
            return RedirectResponse(url=f"/quote?msg=Excel解析失败: {str(e)}&success=0", status_code=303)
    else:
        try:
            text = content.decode('utf-8-sig').strip()
        except UnicodeDecodeError:
            text = content.decode('gbk', errors='replace').strip()
        f = io.StringIO(text)
        reader = csv.reader(f)
        rows_data = list(reader)

    success_count, errors = batch_import_quote_from_rows(rows_data)
    err_msg = ""
    if errors:
        err_msg = "&msg=" + urllib.parse.quote(errors[0])
    return RedirectResponse(url=f"/quote?import_success={success_count}&errors={len(errors)}{err_msg}", status_code=303)


# API 端点
api_router = APIRouter(tags=["quote-api"])


@api_router.post("/api/quote/update")
async def quote_update_api(quote_id: str = Form(...), field: str = Form(...), value: str = Form(default=""), current_user: dict = Depends(login_required)):
    """更新询价API"""
    if current_user['rule'] not in ['3', '0']:
        return {"success": False, "message": "无修改权限"}

    allowed_fields = ['cli_id', 'inquiry_mpn', 'quoted_mpn', 'inquiry_brand', 'inquiry_qty', 'actual_qty', 'target_price_rmb', 'cost_price_rmb', 'date_code', 'delivery_date', 'status', 'is_transferred', 'remark']
    if field not in allowed_fields:
        return {"success": False, "message": f"非法字段: {field}"}

    if field in ['inquiry_qty', 'actual_qty', 'target_price_rmb', 'cost_price_rmb']:
        try:
            val = float(value) if 'price' in field else int(value)
            success, msg = update_quote(quote_id, {field: val})
            return {"success": success, "message": msg}
        except:
            return {"success": False, "message": "必须是数字"}

    success, msg = update_quote(quote_id, {field: value})
    return {"success": success, "message": msg}


@api_router.post("/api/quote/delete")
async def quote_delete_api(quote_id: str = Form(...), current_user: dict = Depends(login_required)):
    """删除询价API"""
    if current_user['rule'] != '3':
        return {"success": False, "message": "仅管理员可删除"}
    success, msg = delete_quote(quote_id)
    return {"success": success, "message": msg}


@api_router.post("/api/quote/batch_delete")
async def quote_batch_delete_api(request: Request, current_user: dict = Depends(login_required)):
    """批量删除询价"""
    if current_user['rule'] != '3':
        return {"success": False, "message": "仅管理员可删除"}
    data = await request.json()
    ids = data.get("ids", [])
    success, msg = batch_delete_quote(ids)
    return {"success": success, "message": msg}


@api_router.post("/api/quote/batch_copy")
async def quote_batch_copy_api(request: Request, current_user: dict = Depends(login_required)):
    """批量复制询价"""
    if current_user['rule'] not in ['3', '0']:
        return {"success": False, "message": "无权限复制需求"}
    data = await request.json()
    ids = data.get("ids", [])
    success, msg = batch_copy_quote(ids)
    return {"success": success, "message": msg}


@api_router.get("/api/quote/info")
async def get_quote_info_api(id: str, current_user: dict = Depends(login_required)):
    """获取询价详情API"""
    with get_db_connection() as conn:
        row = conn.execute("SELECT q.*, c.cli_name FROM uni_quote q LEFT JOIN uni_cli c ON q.cli_id = c.cli_id WHERE q.quote_id = ?", (id,)).fetchone()
        if row:
            return {"success": True, "data": dict(row)}
        return {"success": False, "message": "未找到"}