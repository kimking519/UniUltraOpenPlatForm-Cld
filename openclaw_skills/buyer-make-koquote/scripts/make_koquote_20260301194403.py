"""
buyer-make-koquote: 根据 Excel 模板生成韩文报价单（견적서）
用法:
  python make_koquote_20260301194403.py --offer_ids "b00015"
  python make_koquote_20260301194403.py --offer_ids "b00015,b00016,b00017"

创建时间: 2026-03-01 19:44:03
"""

import sqlite3
import sys
import os
import json
import argparse
from datetime import datetime
from copy import copy

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("错误: 请先安装 openpyxl → pip install openpyxl")
    sys.exit(1)


# ============================================================
# 数据库路径加载
# ============================================================

def load_db_path():
    """
    从 config/db_config_20260301194403.json 读取数据库路径。
    优先级: 环境变量 MAKE_KOQUOTE_DB_PATH > config 文件
    """
    env_path = os.environ.get("MAKE_KOQUOTE_DB_PATH", "")
    if env_path and os.path.exists(env_path):
        return env_path

    config_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "config"
    )

    config_path = None
    if os.path.isdir(config_dir):
        for f in sorted(os.listdir(config_dir), reverse=True):
            if f.startswith("db_config") and f.endswith(".json"):
                config_path = os.path.join(config_dir, f)
                break

    if not config_path or not os.path.exists(config_path):
        print("错误: 配置文件不存在，请检查 config/ 目录")
        sys.exit(1)

    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)

    if sys.platform == "win32":
        return config.get("db_path_windows", "")
    else:
        return config.get("db_path_wsl", "")


# ============================================================
# 查询报价数据
# ============================================================

def query_offers(conn, offer_ids):
    """查询报价记录及其关联的客户信息"""
    placeholders = ','.join(['?'] * len(offer_ids))
    sql = f"""
    SELECT o.offer_id, o.inquiry_mpn, o.quoted_mpn, o.inquiry_brand,
           o.date_code, o.quoted_qty, o.price_kwr, o.offer_price_rmb,
           o.delivery_date, o.remark, o.quote_id,
           q.cli_id, c.cli_name, c.cli_full_name
    FROM uni_offer o
    LEFT JOIN uni_quote q ON o.quote_id = q.quote_id
    LEFT JOIN uni_cli c ON q.cli_id = c.cli_id
    WHERE o.offer_id IN ({placeholders})
    ORDER BY o.offer_id
    """
    rows = conn.execute(sql, offer_ids).fetchall()
    return [dict(r) for r in rows]


# ============================================================
# 复制单元格样式
# ============================================================

def copy_cell_style(source_cell, target_cell):
    """复制单元格的样式（字体、边框、填充、对齐）"""
    if source_cell.font:
        target_cell.font = copy(source_cell.font)
    if source_cell.border:
        target_cell.border = copy(source_cell.border)
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format


# ============================================================
# 数据行合并单元格配置
# 模板中每行数据的合并范围: G-H(수량), J-K(납기), L-M(비고)
# ============================================================

DATA_ROW_MERGES = [
    ("G", "H"),   # 수량(EA) 合并 G-H
    ("J", "K"),   # 납기 合并 J-K
    ("L", "M"),   # 비고 合并 L-M
]

# 模板固定行号
TEMPLATE_DATA_ROWS = [17, 18, 19]  # 模板中的3行数据
TEMPLATE_TOTAL_ROW = 20            # 합계행
TEMPLATE_INFO_ROW = 21             # 견적정보
TEMPLATE_NOTE_ROW = 22             # 특기사항
DATA_ROW_HEIGHT = 39.0             # 数据行高度


# ============================================================
# 生成报价单
# ============================================================

def generate_koquote(offers, template_path, output_path, quote_no, cli_name, cli_full_name):
    """
    基于新版模板生成报价单 Excel。

    模板结构：
      R1-R16 = 头部与列头
      R17+ = 数据行
      数据行后 = 合计行、报价信息行、特别事项行
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    now = datetime.now()
    data_count = len(offers)

    # ---- 1. 填写动态头部信息 ----
    # B5固定填写"수신:"，C5动态填入客户公司全名
    ws.cell(5, 2).value = "수신:"
    ws.cell(5, 3).value = cli_full_name or cli_name or ""

    ws.cell(7, 2).value = f"견적번호 : 제 {quote_no}호"
    ws.cell(9, 2).value = f"작성일자 :{now.year}년  {now.month:02d}월 {now.day:02d}일"

    # ---- 2. 准备数据行样式和合并信息（以R17为蓝本） ----
    first_data_row = 17
    template_row = first_data_row

    template_styles = {}
    for col in range(1, 14):
        cell = ws.cell(template_row, col)
        template_styles[col] = {
            "font": copy(cell.font) if cell.font else None,
            "border": copy(cell.border) if cell.border else None,
            "fill": copy(cell.fill) if cell.fill else None,
            "alignment": copy(cell.alignment) if cell.alignment else None,
            "number_format": cell.number_format,
        }

    # ---- 3. 动态调整行数 ----
    # 数据行后的固定行数: 合计(1行) + 报价信息(1行) + 特别事项(1行) = 3行
    FOOTER_ROWS = 3
    # 计算数据行结束后的位置
    last_data_row = first_data_row + data_count - 1
    # 固定行的起始位置
    total_row = last_data_row + 1          # 合计行
    info_row = last_data_row + 2           # 报价信息行
    note_row = last_data_row + 3           # 特别事项行

    # 删除模板中多余的数据行，保留足够的行给数据和固定行
    rows_needed = data_count + FOOTER_ROWS  # 数据行 + 固定行
    default_template_rows = 20  # 模板默认数据行数

    # 先删除多余的空行
    if data_count < default_template_rows:
        ws.delete_rows(last_data_row + 1, default_template_rows - data_count)

    # ---- 4. 写入数据 ----
    for idx, offer in enumerate(offers):
        row = first_data_row + idx

        ws.cell(row, 2).value = str(idx + 1)                       # No.
        ws.cell(row, 3).value = offer.get("inquiry_mpn", "")       # 모델명
        ws.cell(row, 4).value = offer.get("quoted_mpn", "")        # 제공가능한 부품
        ws.cell(row, 5).value = offer.get("inquiry_brand", "")     # 메이커
        ws.cell(row, 6).value = offer.get("date_code", "")         # 생산일자
        ws.cell(row, 7).value = offer.get("quoted_qty", "")        # 수량

        # 단가(KRW): 优先使用price_kwr，如果为空则使用offer_price_rmb
        price_kwr = offer.get("price_kwr")
        if not price_kwr or str(price_kwr).strip() == "":
            price_kwr = offer.get("offer_price_rmb", "")
        ws.cell(row, 9).value = price_kwr                          # 단가(KRW)

        ws.cell(row, 10).value = offer.get("delivery_date", "")    # 납기
        ws.cell(row, 12).value = offer.get("remark", "")           # 비고

    # ---- 5. 添加固定行（合计、报价信息、特别事项） ----
    # 固定行样式
    footer_font = Font(name='맑은 고딕', size=9.0)
    footer_fill = PatternFill(start_color='F3F3F3', end_color='F3F3F3', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 边框样式
    thin_border = Side(style='thin')
    hair_border = Side(style='hair')

    def set_footer_row_border(ws, row, is_last=False):
        """设置固定行的边框样式"""
        # B列: 左实线，右虚线
        ws.cell(row, 2).border = Border(left=thin_border, right=hair_border, top=hair_border, bottom=hair_border)
        # C列: 右虚线
        ws.cell(row, 3).border = Border(right=hair_border, top=hair_border, bottom=hair_border)
        # D-M列: 左虚线，右虚线
        for col in range(4, 13):
            ws.cell(row, col).border = Border(left=hair_border, right=hair_border, top=hair_border, bottom=hair_border)
        # M列: 右实线
        ws.cell(row, 13).border = Border(left=hair_border, right=thin_border, top=hair_border, bottom=hair_border)

        # 如果是最后一行，底部需要实线
        if is_last:
            for col in range(2, 14):
                cell = ws.cell(row, col)
                existing = cell.border
                cell.border = Border(
                    left=existing.left,
                    right=existing.right,
                    top=existing.top,
                    bottom=thin_border
                )

    # 合计行
    ws.cell(total_row, 2).value = "합    계"
    ws.cell(total_row, 2).font = footer_font
    ws.cell(total_row, 2).fill = footer_fill
    ws.cell(total_row, 2).alignment = center_align
    ws.cell(total_row, 3).fill = footer_fill
    ws.merge_cells(f"B{total_row}:C{total_row}")
    for col_start, col_end in DATA_ROW_MERGES:
        ws.merge_cells(f"{col_start}{total_row}:{col_end}{total_row}")
    ws.row_dimensions[total_row].height = 22.5
    set_footer_row_border(ws, total_row, is_last=False)

    # 报价信息行
    ws.cell(info_row, 2).value = "견적정보"
    ws.cell(info_row, 2).font = footer_font
    ws.cell(info_row, 2).fill = footer_fill
    ws.cell(info_row, 2).alignment = center_align
    ws.cell(info_row, 3).fill = footer_fill
    ws.merge_cells(f"B{info_row}:C{info_row}")
    ws.cell(info_row, 4).value = """1) 견적 유효기간 : 발행 후 3일내
2) 결제방법 : PI에 기재된 홍콩계좌로 입금후 메일로 내역서를 보내주시면  작업이 접수됩니다.
3) 수수료: 고객사측에서 송금수수료 부담,당사측에서 입금수수료 부담
4) 입금계좌 : PI를 참조 바랍니다.
5) 배송택배:1~3일항공
6) A/S기한은 1년이며 품질에 문제가 있을시 무조건 교환/환불이 가능합니다. """
    ws.cell(info_row, 4).font = footer_font
    ws.cell(info_row, 4).alignment = left_align
    ws.merge_cells(f"D{info_row}:M{info_row}")
    ws.row_dimensions[info_row].height = 94.0
    set_footer_row_border(ws, info_row, is_last=False)

    # 特别事项行
    ws.cell(note_row, 2).value = "특기사항"
    ws.cell(note_row, 2).font = footer_font
    ws.cell(note_row, 2).fill = footer_fill
    ws.cell(note_row, 2).alignment = center_align
    ws.cell(note_row, 3).fill = footer_fill
    ws.merge_cells(f"B{note_row}:C{note_row}")
    ws.cell(note_row, 4).value = """1.택배는 1일항공으로 보내드리며 보통은 다음 작업일에 한국에 도착합니다.
2.지정한 청관사가 있을시 알려주시면 교체가 가능하겠습니다."""
    ws.cell(note_row, 4).font = footer_font
    ws.cell(note_row, 4).alignment = left_align
    ws.merge_cells(f"D{note_row}:M{note_row}")
    ws.row_dimensions[note_row].height = 64.0
    set_footer_row_border(ws, note_row, is_last=True)

    # ---- 6. 保存 ----
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


# ============================================================
# 主入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="生成韩文报价单 (buyer-make-koquote)"
    )
    parser.add_argument("--offer_ids", required=True,
                        help="报价编号，多个用逗号分隔")
    parser.add_argument("--db_path", default=None,
                        help="数据库路径（覆盖配置文件）")

    args = parser.parse_args()

    # 解析报价编号
    offer_ids = [oid.strip() for oid in args.offer_ids.split(",") if oid.strip()]
    if not offer_ids:
        print("错误: 请提供至少一个报价编号")
        sys.exit(1)

    # 数据库路径
    db_path = args.db_path or load_db_path()
    if not db_path or not os.path.exists(db_path):
        print(f"错误: 数据库文件不存在 {db_path}")
        sys.exit(1)

    # 模板路径
    skill_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_dir = os.path.join(skill_root, "template")
    template_path = None
    if os.path.isdir(template_dir):
        for f in os.listdir(template_dir):
            if f.endswith(".xlsx"):
                template_path = os.path.join(template_dir, f)
                break

    if not template_path or not os.path.exists(template_path):
        print("错误: 模板文件不存在，请检查 template/ 目录")
        sys.exit(1)

    # 查询数据
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    try:
        offers = query_offers(conn, offer_ids)

        if not offers:
            print(f"❌ 报价编号 {offer_ids} 不存在。")
            sys.exit(1)

        # 检查是否找到所有报价
        found_ids = {o["offer_id"] for o in offers}
        missing = set(offer_ids) - found_ids
        if missing:
            print(f"⚠️ 以下报价编号未找到: {', '.join(missing)}")

        # 检查所有报价是否属于同一客户
        cli_names = set()
        cli_full_names = set()
        for o in offers:
            name = o.get("cli_name") or "未知客户"
            cli_names.add(name)
            full_name = o.get("cli_full_name") or name
            cli_full_names.add(full_name)

        if len(cli_names) > 1:
            print(f"❌ 报价属于不同客户 ({', '.join(cli_names)})，无法生成同一份报价单。")
            sys.exit(1)

        cli_name = list(cli_names)[0]
        cli_full_name = list(cli_full_names)[0] if cli_full_names else cli_name

        # 生成输出路径
        now = datetime.now()
        date_dir = now.strftime("%Y%m%d")
        quote_no = now.strftime("%Y%m%d%H%M")

        # 项目根目录
        project_root = os.path.dirname(os.path.dirname(skill_root))
        output_dir = os.path.join(project_root, "Trans", cli_name, date_dir)
        output_filename = f"유니콘_전자부품견적서_{quote_no}.xlsx"
        output_path = os.path.join(output_dir, output_filename)

        # 生成报价单
        result_path = generate_koquote(offers, template_path, output_path, quote_no, cli_name, cli_full_name)

        # 输出结果
        # 显示相对路径
        try:
            rel_path = os.path.relpath(result_path, project_root)
        except:
            rel_path = result_path

        print(f"✅ 报价单生成成功！")
        print(f"   文件路径: {rel_path}")
        print(f"   报价条数: {len(offers)}")
        print(f"   客    户: {cli_name}")
        print(f"   견적번호: {quote_no}")

    except Exception as e:
        print(f"❌ 操作异常: {str(e)}")
        sys.exit(1)

    finally:
        conn.close()


if __name__ == "__main__":
    main()
