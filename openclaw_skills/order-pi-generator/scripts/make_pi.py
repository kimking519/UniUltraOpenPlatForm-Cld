"""
order-pi-generator: 根据销售订单生成 Proforma Invoice (PI) 文件
用法:
  python make_pi.py --order_ids "SO202602241708360223"
  python make_pi.py --order_ids "SO202602241708360223,SO2026022511215330b3"
  python make_pi.py --order_ids "SO202602"  # 模糊匹配

创建时间: 2026-03-06
"""

import sqlite3
import sys
import os
import json
import argparse
from datetime import datetime
from copy import copy

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("错误: 请先安装 openpyxl -> pip install openpyxl")
    sys.exit(1)


# ============================================================
# 配置加载
# ============================================================

def load_config():
    """加载配置文件"""
    config_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "config"
    )

    config_path = None
    if os.path.isdir(config_dir):
        for f in sorted(os.listdir(config_dir), reverse=True):
            if f.endswith(".json"):
                config_path = os.path.join(config_dir, f)
                break

    if not config_path or not os.path.exists(config_path):
        return {
            "db_path_windows": "E:\\WorkPlace\\7_AI_APP\\UniUltraOpenPlatForm\\uni_platform.db",
            "db_path_wsl": "/home/kim/workspace/UniUltraOpenPlatForm/uni_platform.db",
            "output_base_windows": "E:\\1_Business\\1_Auto",
            "output_base_wsl": "/mnt/e/1_Business/1_Auto",
            "template_name": "Proforma_Invoice_TAEJU_UNI2025110502_v2.xlsx"
        }

    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def get_db_path(config):
    """获取数据库路径"""
    env_path = os.environ.get("SALE_PI_DB_PATH", "")
    if env_path and os.path.exists(env_path):
        return env_path
    if sys.platform == "win32":
        return config.get("db_path_windows", "")
    else:
        return config.get("db_path_wsl", "")


def get_output_base(config):
    """获取输出目录基础路径"""
    if sys.platform == "win32":
        return config.get("output_base_windows", "")
    else:
        return config.get("output_base_wsl", "")


def get_exchange_rate_krw(conn):
    """从数据库获取最新的人民币对韩元汇率"""
    try:
        # currency_code = 2 表示韩元，exchange_rate 是 1人民币=多少韩元
        row = conn.execute("""
            SELECT exchange_rate FROM uni_daily
            WHERE currency_code = 2
            ORDER BY record_date DESC LIMIT 1
        """).fetchone()
        if row and row[0]:
            return float(row[0])
    except:
        pass
    # 默认汇率
    return 218.0


# ============================================================
# 查询订单数据
# ============================================================

def query_orders_by_ids(conn, order_ids, exchange_rate_krw):
    """根据订单ID精确查询，并计算KWR价格"""
    placeholders = ','.join(['?'] * len(order_ids))
    sql = f"""
    SELECT
        o.order_id, o.order_no, o.order_date, o.cli_id,
        o.inquiry_mpn, o.inquiry_brand, o.price_rmb, o.price_kwr, o.price_usd,
        o.offer_id,
        c.cli_name, c.cli_name_en, c.contact_name, c.address, c.email, c.phone,
        c.cli_full_name,
        off.quoted_qty, off.date_code, off.delivery_date, off.inquiry_qty
    FROM uni_order o
    LEFT JOIN uni_cli c ON o.cli_id = c.cli_id
    LEFT JOIN uni_offer off ON o.offer_id = off.offer_id
    WHERE o.order_id IN ({placeholders})
    ORDER BY o.order_id
    """
    rows = conn.execute(sql, order_ids).fetchall()
    results = []
    for r in rows:
        d = dict(r)
        # 计算KWR价格：优先使用price_kwr，否则用price_rmb * 汇率
        price_kwr = d.get("price_kwr")
        if not price_kwr or float(price_kwr or 0) == 0:
            price_rmb = d.get("price_rmb")
            if price_rmb and float(price_rmb or 0) > 0:
                price_kwr = round(float(price_rmb) * exchange_rate_krw, 1)
            else:
                price_kwr = ""
        else:
            price_kwr = float(price_kwr)
        d["calculated_price_kwr"] = price_kwr
        results.append(d)
    return results


def query_orders_fuzzy(conn, keyword, exchange_rate_krw):
    """根据订单编号模糊匹配，并计算KWR价格"""
    sql = """
    SELECT
        o.order_id, o.order_no, o.order_date, o.cli_id,
        o.inquiry_mpn, o.inquiry_brand, o.price_rmb, o.price_kwr, o.price_usd,
        o.offer_id,
        c.cli_name, c.cli_name_en, c.contact_name, c.address, c.email, c.phone,
        c.cli_full_name,
        off.quoted_qty, off.date_code, off.delivery_date, off.inquiry_qty
    FROM uni_order o
    LEFT JOIN uni_cli c ON o.cli_id = c.cli_id
    LEFT JOIN uni_offer off ON o.offer_id = off.offer_id
    WHERE o.order_id LIKE ? OR o.order_no LIKE ?
    ORDER BY o.order_id
    """
    pattern = f"%{keyword}%"
    rows = conn.execute(sql, [pattern, pattern]).fetchall()
    results = []
    for r in rows:
        d = dict(r)
        # 计算KWR价格
        price_kwr = d.get("price_kwr")
        if not price_kwr or float(price_kwr or 0) == 0:
            price_rmb = d.get("price_rmb")
            if price_rmb and float(price_rmb or 0) > 0:
                price_kwr = round(float(price_rmb) * exchange_rate_krw, 1)
            else:
                price_kwr = ""
        else:
            price_kwr = float(price_kwr)
        d["calculated_price_kwr"] = price_kwr
        results.append(d)
    return results


# ============================================================
# 生成PI文件
# ============================================================

def generate_pi(orders, template_path, output_path, invoice_no, cli_name):
    """
    基于模板生成PI Excel文件
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    now = datetime.now()
    data_count = len(orders)

    if data_count == 0:
        return None

    first_order = orders[0]

    # ---- 1. 填写头部信息 ----
    ws.cell(8, 4).value = now.strftime("UNI%Y%m%d%H")
    ws.cell(9, 4).value = now.strftime("%Y-%m-%d")

    cli_name_en = first_order.get("cli_name_en", "") or first_order.get("cli_name", "")
    ws.cell(12, 3).value = cli_name_en
    ws.cell(13, 3).value = first_order.get("contact_name", "") or ""
    ws.cell(14, 3).value = first_order.get("email", "") or ""
    ws.cell(15, 3).value = first_order.get("phone", "") or ""
    ws.cell(16, 3).value = first_order.get("address", "") or ""

    # ---- 2. 处理数据行 ----
    header_row = 18
    first_data_row = 19
    template_data_rows = 2
    total_template_row = 21
    footer_start_row = 22  # TOTAL行之后的固定内容起始行

    # 保存数据行模板的样式 (使用第19行作为模板)
    data_row_styles = {}
    for col in range(1, 9):
        cell = ws.cell(first_data_row, col)
        data_row_styles[col] = {
            "font": copy(cell.font) if cell.font else None,
            "border": copy(cell.border) if cell.border else None,
            "fill": copy(cell.fill) if cell.fill else None,
            "alignment": copy(cell.alignment) if cell.alignment else None,
            "number_format": cell.number_format,
        }
    template_row_height = ws.row_dimensions[first_data_row].height or 20.0

    # 保存TOTAL AMOUNT行的样式
    total_row_styles = {}
    for col in range(1, 9):
        cell = ws.cell(total_template_row, col)
        total_row_styles[col] = {
            "font": copy(cell.font) if cell.font else None,
            "border": copy(cell.border) if cell.border else None,
            "fill": copy(cell.fill) if cell.fill else None,
            "alignment": copy(cell.alignment) if cell.alignment else None,
            "number_format": cell.number_format,
        }
    total_row_height = ws.row_dimensions[total_template_row].height or 20.0

    # ============================================================
    # 关键修复：在插入行之前保存固定内容区域的所有信息
    # ============================================================

    # 保存固定内容区域（Row 22及以后）的合并单元格
    footer_merged_ranges = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= footer_start_row:
            footer_merged_ranges.append({
                "range_str": str(merged_range),
                "min_row": merged_range.min_row,
                "min_col": merged_range.min_col,
                "max_row": merged_range.max_row,
                "max_col": merged_range.max_col,
            })

    # 保存固定内容区域的所有单元格值和样式
    footer_cells_data = {}
    for row in range(footer_start_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            # 只保存有内容的单元格
            if cell.value is not None or cell.font or cell.border or cell.fill:
                footer_cells_data[(row, col)] = {
                    "value": cell.value,
                    "font": copy(cell.font) if cell.font else None,
                    "border": copy(cell.border) if cell.border else None,
                    "fill": copy(cell.fill) if cell.fill else None,
                    "alignment": copy(cell.alignment) if cell.alignment else None,
                    "number_format": cell.number_format,
                }

    # 保存行高
    footer_row_heights = {}
    for row in range(footer_start_row, ws.max_row + 1):
        if ws.row_dimensions[row].height:
            footer_row_heights[row] = ws.row_dimensions[row].height

    # 根据数据量调整行
    rows_diff = data_count - template_data_rows
    if rows_diff > 0:
        # 需要插入新行
        ws.insert_rows(total_template_row, rows_diff)
        for i in range(rows_diff):
            new_row = total_template_row + i
            ws.row_dimensions[new_row].height = template_row_height
    elif rows_diff < 0:
        # 需要删除多余行
        rows_to_delete = -rows_diff
        ws.delete_rows(first_data_row + data_count, rows_to_delete)

    # 计算实际的TOTAL行位置
    actual_total_row = first_data_row + data_count
    actual_footer_start_row = actual_total_row + 1

    # ============================================================
    # 关键修复：恢复固定内容区域的数据
    # ============================================================

    # 先取消所有可能存在的合并单元格
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= actual_footer_start_row:
            try:
                ws.unmerge_cells(str(merged_range))
            except:
                pass

    # 恢复单元格值和样式（按新的行号偏移）
    row_offset = rows_diff  # 行偏移量
    for (orig_row, col), cell_data in footer_cells_data.items():
        new_row = orig_row + row_offset
        if new_row >= actual_footer_start_row:
            cell = ws.cell(new_row, col)
            if cell_data["value"] is not None:
                cell.value = cell_data["value"]
            if cell_data["font"]:
                cell.font = cell_data["font"]
            if cell_data["border"]:
                cell.border = cell_data["border"]
            if cell_data["fill"]:
                cell.fill = cell_data["fill"]
            if cell_data["alignment"]:
                cell.alignment = cell_data["alignment"]
            if cell_data["number_format"]:
                cell.number_format = cell_data["number_format"]

    # 恢复行高
    for orig_row, height in footer_row_heights.items():
        new_row = orig_row + row_offset
        ws.row_dimensions[new_row].height = height

    # 恢复合并单元格
    for merged_info in footer_merged_ranges:
        new_min_row = merged_info["min_row"] + row_offset
        new_max_row = merged_info["max_row"] + row_offset
        # 跳过会覆盖数据行的合并单元格
        if new_min_row <= actual_total_row - 1 and new_max_row >= first_data_row:
            continue
        try:
            ws.merge_cells(
                start_row=new_min_row,
                start_column=merged_info["min_col"],
                end_row=new_max_row,
                end_column=merged_info["max_col"]
            )
        except:
            pass

    # 取消数据行的合并单元格（如果有）
    merged_to_remove = []
    for merged_range in list(ws.merged_cells.ranges):
        min_row = merged_range.min_row
        if first_data_row <= min_row <= actual_total_row - 1:
            merged_to_remove.append(merged_range)
    for merged_range in merged_to_remove:
        try:
            ws.unmerge_cells(str(merged_range))
        except:
            pass

    # 写入数据
    for idx, order in enumerate(orders):
        row = first_data_row + idx

        # A列 = # (序号)
        ws.cell(row, 1).value = idx + 1
        # B列 = Part No. (型号)
        ws.cell(row, 2).value = order.get("inquiry_mpn", "") or ""
        # C列 = Manufacturer (品牌)
        ws.cell(row, 3).value = order.get("inquiry_brand", "") or ""
        # D列 = D/C (批号)
        ws.cell(row, 4).value = order.get("date_code", "") or ""
        # E列 = Qty (数量) - 整数格式带千位分隔符
        qty = order.get("quoted_qty") or order.get("inquiry_qty") or ""
        ws.cell(row, 5).value = qty
        if qty:
            ws.cell(row, 5).number_format = '#,##0'
        # F列 = L/T (货期)
        ws.cell(row, 6).value = order.get("delivery_date", "") or ""
        # G列 = Unit Price (KRW) - 整数格式带千位分隔符
        price_kwr = order.get("calculated_price_kwr", "") or ""
        ws.cell(row, 7).value = price_kwr
        if price_kwr:
            ws.cell(row, 7).number_format = '#,##0'
        # H列 = Total Amount = G * E
        if qty and price_kwr:
            ws.cell(row, 8).value = f"=G{row}*E{row}"
            ws.cell(row, 8).number_format = '#,##0'
        else:
            ws.cell(row, 8).value = ""

        # 应用样式
        for col in range(1, 9):
            cell = ws.cell(row, col)
            style = data_row_styles.get(col, {})
            if style.get("font"):
                cell.font = style["font"]
            if style.get("border"):
                cell.border = style["border"]
            if style.get("fill"):
                cell.fill = style["fill"]
            if style.get("alignment"):
                cell.alignment = style["alignment"]

    # ---- 3. 更新TOTAL AMOUNT行 ----
    last_data_row = actual_total_row - 1

    # 取消TOTAL行的合并单元格
    merged_to_unmerge = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row == actual_total_row:
            merged_to_unmerge.append(merged_range)
    for merged_range in merged_to_unmerge:
        try:
            ws.unmerge_cells(str(merged_range))
        except:
            pass

    # 设置TOTAL AMOUNT值 - 改为Total Amount:
    ws.cell(actual_total_row, 1).value = "Total Amount:"
    ws.cell(actual_total_row, 8).value = f"=SUM(H{first_data_row}:H{last_data_row})"
    ws.cell(actual_total_row, 8).number_format = '#,##0'  # 千位分隔符

    # 重新合并A到G列
    try:
        ws.merge_cells(f"A{actual_total_row}:G{actual_total_row}")
    except:
        pass

    # 应用TOTAL行样式
    for col in range(1, 9):
        try:
            cell = ws.cell(actual_total_row, col)
            style = total_row_styles.get(col, {})
            if style.get("font"):
                cell.font = style["font"]
            if style.get("border"):
                cell.border = style["border"]
            if style.get("fill"):
                cell.fill = style["fill"]
            if style.get("alignment"):
                cell.alignment = style["alignment"]
        except:
            pass

    # 覆盖样式：标题白色右对齐，值红色居中
    ws.cell(actual_total_row, 1).font = Font(bold=True, color="FFFFFF")  # 白色字体
    ws.cell(actual_total_row, 1).alignment = Alignment(horizontal="right")  # 右对齐
    ws.cell(actual_total_row, 8).font = Font(bold=True, color="FF0000")  # 红色字体
    ws.cell(actual_total_row, 8).alignment = Alignment(horizontal="center")  # 居中对齐

    # ---- 4. 调整印章图片位置 ----
    # TERMS & CONDITIONS 行在 TOTAL 行 + 2 的位置
    terms_row = actual_total_row + 2
    # 印章图片放在 TERMS & CONDITIONS 行下方
    stamp_row = terms_row + 1

    for img in ws._images:
        # 保存原始尺寸
        orig_width = img.width
        orig_height = img.height

        anchor = img.anchor

        # 更新锚点位置
        if hasattr(anchor, '_from'):
            # 计算原始行跨度（如果是 TwoCellAnchor）
            if hasattr(anchor, 'to'):
                orig_row_span = anchor.to.row - anchor._from.row
                # 更新起始和结束位置
                anchor._from.row = stamp_row
                anchor.to.row = stamp_row + orig_row_span
            else:
                # OneCellAnchor 只更新起始位置
                anchor._from.row = stamp_row

        # 确保图片保持原始尺寸
        img.width = orig_width
        img.height = orig_height

    # ---- 5. 更新打印区域 ----
    # 动态更新打印区域，确保所有内容都能打印
    ws.print_area = f"$A$1:$H${ws.max_row}"

    # ---- 6. 保存文件 ----
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


# ============================================================
# 主入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="生成 Proforma Invoice (order-pi-generator)")
    parser.add_argument("--order_ids", required=True, help="订单编号，多个用逗号分隔，支持模糊匹配")
    parser.add_argument("--db_path", default=None, help="数据库路径（覆盖配置文件）")
    parser.add_argument("--output_dir", default=None, help="输出目录（覆盖配置文件）")

    args = parser.parse_args()

    order_ids = [oid.strip() for oid in args.order_ids.split(",") if oid.strip()]
    if not order_ids:
        print("错误: 请提供至少一个订单编号")
        sys.exit(1)

    config = load_config()
    db_path = args.db_path or get_db_path(config)
    if not db_path or not os.path.exists(db_path):
        print(f"错误: 数据库文件不存在 {db_path}")
        sys.exit(1)

    output_base = args.output_dir or get_output_base(config)
    if not output_base:
        output_base = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
            "Trans"
        )

    skill_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_dir = os.path.join(skill_root, "template")
    template_path = None

    template_name = config.get("template_name", "")
    if template_name:
        template_path = os.path.join(template_dir, template_name)
        if not os.path.exists(template_path):
            template_path = None

    if not template_path and os.path.isdir(template_dir):
        for f in os.listdir(template_dir):
            if f.endswith(".xlsx"):
                template_path = os.path.join(template_dir, f)
                break

    if not template_path or not os.path.exists(template_path):
        print("错误: 模板文件不存在")
        sys.exit(1)

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    try:
        # 获取汇率
        exchange_rate_krw = get_exchange_rate_krw(conn)

        orders = query_orders_by_ids(conn, order_ids, exchange_rate_krw)

        if not orders and len(order_ids) == 1:
            orders = query_orders_fuzzy(conn, order_ids[0], exchange_rate_krw)
            if orders:
                print(f"提示: 使用模糊匹配 '{order_ids[0]}' 找到 {len(orders)} 条订单")

        if not orders:
            print(f"错误: 订单编号 {order_ids} 不存在")
            sys.exit(1)

        if len(order_ids) > 1 or (len(order_ids) == 1 and orders and orders[0]['order_id'] in order_ids):
            found_ids = {o["order_id"] for o in orders}
            missing = set(order_ids) - found_ids
            if missing:
                print(f"警告: 以下订单编号未找到: {', '.join(missing)}")

        cli_names = set(o.get("cli_name") or "未知客户" for o in orders)
        if len(cli_names) > 1:
            print(f"错误: 订单属于不同客户 ({', '.join(cli_names)})，无法生成同一份PI")
            sys.exit(1)

        cli_name = list(cli_names)[0]
        now = datetime.now()
        date_dir = now.strftime("%Y%m%d")
        invoice_no = now.strftime("UNI%Y%m%d%H%M%S")

        output_dir = os.path.join(output_base, cli_name, date_dir)
        output_filename = f"Proforma Invoice_{cli_name}_{invoice_no}.xlsx"
        output_path = os.path.join(output_dir, output_filename)

        result_path = generate_pi(orders, template_path, output_path, invoice_no, cli_name)

        if result_path:
            print(f"成功 PI生成成功！")
            print(f"文件路径: {result_path}")
            print(f"订单条数: {len(orders)}")
            print(f"客户: {cli_name}")
            print(f"Invoice No.: {now.strftime('UNI%Y%m%d%H')}")
        else:
            print("错误: PI生成失败")
            sys.exit(1)

    except Exception as e:
        print(f"错误: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        conn.close()


if __name__ == "__main__":
    main()