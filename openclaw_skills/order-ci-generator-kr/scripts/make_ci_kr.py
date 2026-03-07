"""
order-ci-generator-kr: 根据销售订单生成 Commercial Invoice (CI) 文件
用法:
  python make_ci_kr.py --order_ids "SO202602241708360223"
  python make_ci_kr.py --order_ids "SO202602241708360223,SO2026022511215330b3"
  python make_ci_kr.py --order_ids "SO202602"  # 模糊匹配

创建时间: 2026-03-07
"""

import sqlite3
import sys
import os
import json
import argparse
from datetime import datetime
from copy import copy

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("错误: 请先安装 openpyxl -> pip install openpyxl")
    sys.exit(1)

# 尝试导入xlrd以支持WPS格式
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


# ============================================================
# WPS格式转换支持
# ============================================================

def convert_wps_to_xlsx(template_path):
    """将WPS格式的xlsx文件转换为标准xlsx格式"""
    if not XLRD_AVAILABLE:
        return None

    try:
        import tempfile
        import shutil

        # 读取WPS文件
        wb_old = xlrd.open_workbook(template_path)
        ws_old = wb_old.sheet_by_index(0)

        # 创建新的openpyxl工作簿
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        ws_new.title = ws_old.name

        # 复制单元格数据
        for row_idx in range(ws_old.nrows):
            for col_idx in range(ws_old.ncols):
                try:
                    cell = ws_old.cell_value(row_idx, col_idx)
                    # 处理日期格式
                    if isinstance(cell, float) and ws_old.cell_type(row_idx, col_idx) == 3:
                        # Excel日期格式
                        from datetime import datetime as dt
                        cell = dt(*xlrd.xldate_as_tuple(cell, wb_old.datemode))

                    if cell:
                        ws_new.cell(row_idx + 1, col_idx + 1).value = cell
                except:
                    pass

        # 复制行高
        for row_idx in range(ws_old.nrows):
            try:
                height = ws_old.row_info_list[row_idx].height
                if height:
                    ws_new.row_dimensions[row_idx + 1].height = height / 20  # xlrd height is in twips
            except:
                pass

        # 保存到临时文件
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, "ci_template_converted.xlsx")
        wb_new.save(temp_path)

        return temp_path
    except Exception as e:
        print(f"警告: WPS格式转换失败: {e}")
        return None


def load_workbook_safe(template_path):
    """安全加载工作簿，支持WPS格式"""
    try:
        # 先尝试标准方式
        return openpyxl.load_workbook(template_path)
    except Exception as e:
        # 如果失败，尝试转换WPS格式
        if XLRD_AVAILABLE:
            converted_path = convert_wps_to_xlsx(template_path)
            if converted_path and os.path.exists(converted_path):
                try:
                    return openpyxl.load_workbook(converted_path)
                except:
                    pass
        raise e


# ============================================================
# 配置加载
# ============================================================

def load_config():
    """加载配置文件"""
    config_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "config"
    )

    config_path = None
    if os.path.isdir(config_dir):
        for f in sorted(os.listdir(config_dir), reverse=True):
            if f.endswith(".json"):
                config_path = os.path.join(config_dir, f)
                break

    if not config_path or not os.path.exists(config_path):
        return {
            "db_path_windows": "E:\\WorkPlace\\7_AI_APP\\UniUltraOpenPlatForm\\uni_platform.db",
            "db_path_wsl": "/home/kim/workspace/UniUltraOpenPlatForm/uni_platform.db",
            "output_base_windows": "E:\\1_Business\\1_Auto",
            "output_base_wsl": "/mnt/e/1_Business/1_Auto",
            "template_name": "COMMERCIAL INVOICE_TAEJU_UNI2026012101_C.xlsx"
        }

    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def get_db_path(config):
    """获取数据库路径"""
    env_path = os.environ.get("SALE_CI_DB_PATH", "")
    if env_path and os.path.exists(env_path):
        return env_path
    if sys.platform == "win32":
        return config.get("db_path_windows", "")
    else:
        return config.get("db_path_wsl", "")


def get_output_base(config):
    """获取输出目录基础路径"""
    if sys.platform == "win32":
        return config.get("output_base_windows", "")
    else:
        return config.get("output_base_wsl", "")


# ============================================================
# 查询订单数据
# ============================================================

def get_exchange_rate_krw(conn):
    """从数据库获取最新的人民币对韩元汇率"""
    try:
        # currency_code = 2 表示韩元，exchange_rate 是 1人民币=多少韩元
        row = conn.execute("""
            SELECT exchange_rate FROM uni_daily
            WHERE currency_code = 2
            ORDER BY record_date DESC LIMIT 1
        """).fetchone()
        if row and row[0]:
            return float(row[0])
    except:
        pass
    # 默认汇率
    return 218.0


def calculate_price_kwr(order, exchange_rate_krw):
    """计算韩元价格"""
    price_kwr = order.get("price_kwr")
    if not price_kwr or float(price_kwr or 0) == 0:
        price_rmb = order.get("price_rmb")
        if price_rmb and float(price_rmb or 0) > 0:
            price_kwr = round(float(price_rmb) * exchange_rate_krw, 1)
        else:
            price_kwr = 0
    else:
        price_kwr = float(price_kwr)
    return price_kwr


def query_orders_by_ids(conn, order_ids, exchange_rate_krw):
    """根据订单ID精确查询"""
    placeholders = ','.join(['?'] * len(order_ids))
    sql = f"""
    SELECT
        o.order_id, o.order_no, o.order_date, o.cli_id,
        o.inquiry_mpn, o.inquiry_brand, o.price_rmb, o.price_kwr, o.price_usd,
        o.offer_id,
        c.cli_name, c.cli_name_en, c.contact_name, c.address, c.email, c.phone,
        c.cli_full_name, c.region,
        off.quoted_qty, off.date_code, off.delivery_date, off.inquiry_qty
    FROM uni_order o
    LEFT JOIN uni_cli c ON o.cli_id = c.cli_id
    LEFT JOIN uni_offer off ON o.offer_id = off.offer_id
    WHERE o.order_id IN ({placeholders})
    ORDER BY o.order_id
    """
    rows = conn.execute(sql, order_ids).fetchall()
    results = []
    for r in rows:
        d = dict(r)
        d["calculated_price_kwr"] = calculate_price_kwr(d, exchange_rate_krw)
        results.append(d)
    return results


def query_orders_fuzzy(conn, keyword, exchange_rate_krw):
    """根据订单编号模糊匹配"""
    sql = """
    SELECT
        o.order_id, o.order_no, o.order_date, o.cli_id,
        o.inquiry_mpn, o.inquiry_brand, o.price_rmb, o.price_kwr, o.price_usd,
        o.offer_id,
        c.cli_name, c.cli_name_en, c.contact_name, c.address, c.email, c.phone,
        c.cli_full_name, c.region,
        off.quoted_qty, off.date_code, off.delivery_date, off.inquiry_qty
    FROM uni_order o
    LEFT JOIN uni_cli c ON o.cli_id = c.cli_id
    LEFT JOIN uni_offer off ON o.offer_id = off.offer_id
    WHERE o.order_id LIKE ? OR o.order_no LIKE ?
    ORDER BY o.order_id
    """
    pattern = f"%{keyword}%"
    rows = conn.execute(sql, [pattern, pattern]).fetchall()
    results = []
    for r in rows:
        d = dict(r)
        d["calculated_price_kwr"] = calculate_price_kwr(d, exchange_rate_krw)
        results.append(d)
    return results


# ============================================================
# 生成CI文件
# ============================================================

def generate_ci(orders, template_path, output_path, cli_name):
    """
    基于模板生成CI Excel文件
    """
    # 直接使用openpyxl加载标准xlsx模板
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    now = datetime.now()
    data_count = len(orders)

    if data_count == 0:
        return None

    first_order = orders[0]

    # ---- 1. 填写头部信息 ----
    # B2 = UNI + 日期(yyyymmddhh)
    ws.cell(2, 2).value = now.strftime("UNI%Y%m%d%H")
    # F2 = 日期 YYYY-MM-DD
    ws.cell(2, 6).value = now.strftime("%Y-%m-%d")

    # B9 = 公司英文名
    cli_name_en = first_order.get("cli_name_en", "") or first_order.get("cli_name", "")
    ws.cell(9, 2).value = cli_name_en
    # F9 = 联系人
    ws.cell(9, 6).value = first_order.get("contact_name", "") or ""
    # B10 = 国家
    ws.cell(10, 2).value = first_order.get("region", "") or "韩国"
    # F10 = 电话
    ws.cell(10, 6).value = first_order.get("phone", "") or ""
    # B11 = 地址
    ws.cell(11, 2).value = first_order.get("address", "") or ""

    # ---- 2. 处理数据行 ----
    # 数据从第13行开始
    first_data_row = 13
    template_data_rows = 3  # 模板中有3行示例数据
    total_template_row = 16  # Total Qty行

    # 保存数据行模板的样式 (使用第13行作为模板)
    data_row_styles = {}
    for col in range(1, 8):
        cell = ws.cell(first_data_row, col)
        data_row_styles[col] = {
            "font": copy(cell.font) if cell.font else None,
            "border": copy(cell.border) if cell.border else None,
            "fill": copy(cell.fill) if cell.fill else None,
            "alignment": copy(cell.alignment) if cell.alignment else None,
            "number_format": cell.number_format,
        }
    template_row_height = ws.row_dimensions[first_data_row].height or 22.0

    # 保存Total行的样式
    total_row_styles = {}
    for col in range(1, 8):
        cell = ws.cell(total_template_row, col)
        total_row_styles[col] = {
            "font": copy(cell.font) if cell.font else None,
            "border": copy(cell.border) if cell.border else None,
            "fill": copy(cell.fill) if cell.fill else None,
            "alignment": copy(cell.alignment) if cell.alignment else None,
            "number_format": cell.number_format,
        }

    # 保存Total Amount行(Row 17)的样式
    total_amount_row_styles = {}
    for col in range(1, 8):
        cell = ws.cell(17, col)
        total_amount_row_styles[col] = {
            "font": copy(cell.font) if cell.font else None,
            "border": copy(cell.border) if cell.border else None,
            "fill": copy(cell.fill) if cell.fill else None,
            "alignment": copy(cell.alignment) if cell.alignment else None,
            "number_format": cell.number_format,
        }

    # 根据数据量调整行
    rows_diff = data_count - template_data_rows
    if rows_diff > 0:
        # 需要插入新行
        ws.insert_rows(total_template_row, rows_diff)
        for i in range(rows_diff):
            new_row = total_template_row + i
            ws.row_dimensions[new_row].height = template_row_height
    elif rows_diff < 0:
        # 需要删除多余行
        rows_to_delete = -rows_diff
        ws.delete_rows(first_data_row + data_count, rows_to_delete)

    # 计算实际的Total行位置
    actual_total_qty_row = first_data_row + data_count
    actual_total_amount_row = first_data_row + data_count + 1

    # 写入数据行
    total_qty = 0
    total_amount = 0

    for idx, order in enumerate(orders):
        row = first_data_row + idx

        # A列 = # (序号)
        ws.cell(row, 1).value = idx + 1

        # B列 = Part No (型号)
        ws.cell(row, 2).value = order.get("inquiry_mpn", "") or ""

        # C列 = DESCRIPTION OF GOODS (固定值: 集成电路/IC)
        ws.cell(row, 3).value = "集成电路/IC"

        # D列 = HS Code (固定值: 8542399000)
        ws.cell(row, 4).value = "8542399000"

        # E列 = Qty (数量)
        qty = order.get("quoted_qty") or order.get("inquiry_qty") or 0
        try:
            qty = int(qty)
        except:
            qty = 0
        ws.cell(row, 5).value = qty
        if qty:
            ws.cell(row, 5).number_format = '#,##0'
        total_qty += qty

        # F列 = Unit Price (KRW)
        price_kwr = order.get("calculated_price_kwr", 0)
        try:
            price_kwr = float(price_kwr)
        except:
            price_kwr = 0
        ws.cell(row, 6).value = price_kwr
        if price_kwr:
            ws.cell(row, 6).number_format = '#,##0'

        # G列 = Total (price_kwr * qty)
        total = price_kwr * qty
        ws.cell(row, 7).value = total if total else 0
        if total:
            ws.cell(row, 7).number_format = '#,##0'
        total_amount += total

        # 应用数据行样式
        for col in range(1, 8):
            cell = ws.cell(row, col)
            style = data_row_styles.get(col, {})
            if style.get("font"):
                cell.font = style["font"]
            if style.get("border"):
                cell.border = style["border"]
            if style.get("fill"):
                cell.fill = style["fill"]
            if style.get("alignment"):
                cell.alignment = style["alignment"]

    # ---- 3. 更新TOTAL行 ----
    # 取消Total行的合并单元格
    merged_to_unmerge = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row == actual_total_qty_row or merged_range.min_row == actual_total_amount_row:
            merged_to_unmerge.append(merged_range)
    for merged_range in merged_to_unmerge:
        try:
            ws.unmerge_cells(str(merged_range))
        except:
            pass

    # Total Qty行
    # D列 = "Total Qty:"
    ws.cell(actual_total_qty_row, 4).value = "Total Qty:"
    ws.cell(actual_total_qty_row, 4).font = Font(name='Arial', size=10, bold=True)
    ws.cell(actual_total_qty_row, 4).alignment = Alignment(horizontal='right', vertical='center')

    # E列 = 总数量
    ws.cell(actual_total_qty_row, 5).value = total_qty
    ws.cell(actual_total_qty_row, 5).number_format = '#,##0'
    ws.cell(actual_total_qty_row, 5).font = Font(name='Arial', size=10, bold=True)
    ws.cell(actual_total_qty_row, 5).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(actual_total_qty_row, 5).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    # 合并A16:C16
    try:
        ws.merge_cells(f"A{actual_total_qty_row}:C{actual_total_qty_row}")
    except:
        pass

    # Total Amount行
    # A列 = "Total invoice amount:"
    ws.cell(actual_total_amount_row, 1).value = "Total invoice amount:"
    ws.cell(actual_total_amount_row, 1).font = Font(name='Arial', size=10, bold=True)
    ws.cell(actual_total_amount_row, 1).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(actual_total_amount_row, 1).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    # G列 = 总金额
    ws.cell(actual_total_amount_row, 7).value = total_amount
    ws.cell(actual_total_amount_row, 7).number_format = '#,##0'
    ws.cell(actual_total_amount_row, 7).font = Font(name='Arial', size=10, bold=True, color='FF0000')
    ws.cell(actual_total_amount_row, 7).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(actual_total_amount_row, 7).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    # 合并A17:F17
    try:
        ws.merge_cells(f"A{actual_total_amount_row}:F{actual_total_amount_row}")
    except:
        pass

    # 应用边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for col in range(1, 8):
        ws.cell(actual_total_qty_row, col).border = thin_border
        ws.cell(actual_total_amount_row, col).border = thin_border

    # ---- 4. 保留模板图片（位置不变） ----
    # CI模板不需要动态插入/删除行，图片保持原位置
    # 如需调整图片大小，可在此处修改
    for img in ws._images:
        # 保持图片原始尺寸
        pass

    # ---- 5. 更新打印区域 ----
    ws.print_area = f"$A$1:$G${ws.max_row}"

    # ---- 6. 保存文件 ----
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


# ============================================================
# 主入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="生成 Commercial Invoice (order-ci-generator-kr)")
    parser.add_argument("--order_ids", required=True, help="订单编号，多个用逗号分隔，支持模糊匹配")
    parser.add_argument("--db_path", default=None, help="数据库路径（覆盖配置文件）")
    parser.add_argument("--output_dir", default=None, help="输出目录（覆盖配置文件）")

    args = parser.parse_args()

    order_ids = [oid.strip() for oid in args.order_ids.split(",") if oid.strip()]
    if not order_ids:
        print("错误: 请提供至少一个订单编号")
        sys.exit(1)

    config = load_config()
    db_path = args.db_path or get_db_path(config)
    if not db_path or not os.path.exists(db_path):
        print(f"错误: 数据库文件不存在 {db_path}")
        sys.exit(1)

    output_base = args.output_dir or get_output_base(config)
    if not output_base:
        output_base = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
            "Trans"
        )

    skill_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_dir = os.path.join(skill_root, "template")
    template_path = None

    template_name = config.get("template_name", "")
    if template_name:
        template_path = os.path.join(template_dir, template_name)
        if not os.path.exists(template_path):
            template_path = None

    if not template_path and os.path.isdir(template_dir):
        for f in os.listdir(template_dir):
            if f.endswith(".xlsx"):
                template_path = os.path.join(template_dir, f)
                break

    if not template_path or not os.path.exists(template_path):
        print("错误: 模板文件不存在")
        sys.exit(1)

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    try:
        # 获取汇率
        exchange_rate_krw = get_exchange_rate_krw(conn)

        orders = query_orders_by_ids(conn, order_ids, exchange_rate_krw)

        if not orders and len(order_ids) == 1:
            orders = query_orders_fuzzy(conn, order_ids[0], exchange_rate_krw)
            if orders:
                print(f"提示: 使用模糊匹配 '{order_ids[0]}' 找到 {len(orders)} 条订单")

        if not orders:
            print(f"错误: 订单编号 {order_ids} 不存在")
            sys.exit(1)

        if len(order_ids) > 1 or (len(order_ids) == 1 and orders and orders[0]['order_id'] in order_ids):
            found_ids = {o["order_id"] for o in orders}
            missing = set(order_ids) - found_ids
            if missing:
                print(f"警告: 以下订单编号未找到: {', '.join(missing)}")

        cli_names = set(o.get("cli_name") or "未知客户" for o in orders)
        if len(cli_names) > 1:
            print(f"错误: 订单属于不同客户 ({', '.join(cli_names)})，无法生成同一份CI")
            sys.exit(1)

        cli_name = list(cli_names)[0]
        now = datetime.now()
        date_dir = now.strftime("%Y%m%d")
        invoice_no = now.strftime("UNI%Y%m%d%H%M%S")

        output_dir = os.path.join(output_base, cli_name, date_dir)
        output_filename = f"COMMERCIAL INVOICE_{cli_name}_{invoice_no}.xlsx"
        output_path = os.path.join(output_dir, output_filename)

        result_path = generate_ci(orders, template_path, output_path, cli_name)

        if result_path:
            print(f"成功 CI生成成功！")
            print(f"文件路径: {result_path}")
            print(f"订单条数: {len(orders)}")
            print(f"客户: {cli_name}")
            print(f"Invoice No.: {now.strftime('UNI%Y%m%d%H')}")
        else:
            print("错误: CI生成失败")
            sys.exit(1)

    except Exception as e:
        print(f"错误: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        conn.close()


if __name__ == "__main__":
    main()